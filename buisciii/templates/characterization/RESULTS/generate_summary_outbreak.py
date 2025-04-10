import os
import glob
import pandas as pd
from openpyxl import load_workbook

# ------------------------------------------------------
# File Paths
# ------------------------------------------------------

# Paths to input files and directories
xlsx_template = "summary_outbreak_template.xlsx"
samples_file = "../ANALYSIS/samples_id.txt"
tsv_file = glob.glob("../ANALYSIS/*_CHARACTERIZATION/99-stats/ariba_mlst_full.tsv")
csv_file = glob.glob("../ANALYSIS/*_ASSEMBLY/99-stats/kmerfinder_summary.csv")
mapping_file = glob.glob("../ANALYSIS/*_SNIPPY/99-stats/mapping_stats_summary.txt")
wgs_metrics_file = glob.glob(
    "../ANALYSIS/*_SNIPPY/99-stats/wgs_metrics_all_filtered.txt"
)
variants_stats_file = glob.glob("../ANALYSIS/*_SNIPPY/99-stats/variants_stats.txt")
quast_report_file = glob.glob(
    "../ANALYSIS/*_ASSEMBLY/03-assembly/quast/global_report/transposed_report.tsv"
)
quast_dir = glob.glob("../ANALYSIS/*_ASSEMBLY/03-assembly/quast/per_reference_reports/")
virulence_file = glob.glob(
    "../ANALYSIS/*_CHARACTERIZATION/99-stats/ariba_vfdb_full.csv"
)
card_file = glob.glob("../ANALYSIS/*_CHARACTERIZATION/99-stats/ariba_card.csv")
amrfinder_dir = glob.glob("../ANALYSIS/*_CHARACTERIZATION/03-amrfinderplus")
plasmid_file = glob.glob(
    "../ANALYSIS/*_PLASMIDID/NO_GROUP/NO_GROUP_final_results_per_sample.tab"
)
mlva_file = glob.glob(
    "../ANALYSIS/*_CHARACTERIZATION/05-mlva/MLVA_output/MLVA_analysis_assemblies.csv"
)


def get_first_match(file_list):
    return file_list[0] if file_list else None


tsv_file = get_first_match(tsv_file)
csv_file = get_first_match(csv_file)
mapping_file = get_first_match(mapping_file)
wgs_metrics_file = get_first_match(wgs_metrics_file)
variants_stats_file = get_first_match(variants_stats_file)
quast_report_file = get_first_match(quast_report_file)
quast_dir = get_first_match(quast_dir)
virulence_file = get_first_match(virulence_file)
card_file = get_first_match(card_file)
amrfinder_dir = get_first_match(amrfinder_dir)
plasmid_file = get_first_match(plasmid_file)
mlva_file = get_first_match(mlva_file)

# ------------------------------------------------------
# File Reading Functions
# ------------------------------------------------------


def read_samples(file_path):
    """
    Reads sample identifiers from a file and maps them to row numbers for the Excel sheet.

    Args:
        file_path (str): Path to the file containing sample identifiers.

    Returns:
        dict: A dictionary where keys are sample identifiers and values are their corresponding row numbers.
    """
    samples = {}
    with open(file_path, "r") as f:
        for index, line in enumerate(f, start=3):  # Start at row 3 in Excel
            sample_id = line.strip()
            samples[sample_id] = index
    return samples


def read_ariba_mlst(tsv_file):
    """
    Processes the MLST TSV file and extracts relevant sequence typing data.

    Args:
        tsv_file (str): Path to the MLST results file.

    Returns:
        tuple: A dictionary mapping sample IDs to their sequence types (ST) and gene sequences,
               and a header string for the gene columns.
    """
    df_mlst = pd.read_csv(tsv_file, sep="\t")
    mlst_columns = df_mlst.columns[2:].tolist()
    mlst_header = "/".join(mlst_columns)
    mlst_dict = {
        str(row["sample_id"]): {
            "ST": row["ST"],
            "genes": "/".join(map(str, row[mlst_columns].values)),
        }
        for _, row in df_mlst.iterrows()
    }
    return mlst_dict, mlst_header


def read_kmerfinder(csv_file):
    """
    Reads a CSV file and extracts specific columns for each sample.

    Args:
        csv_file (str): Path to the CSV file.

    Returns:
        dict: A dictionary where keys are sample names and values are dictionaries containing
              selected columns ('colE', 'colF', 'colG').
    """
    try:
        df_kmerfinder = pd.read_csv(csv_file)

        # Ensure the expected columns exist
        expected_columns = [
            "sample_name",
            df_kmerfinder.columns[1],
            df_kmerfinder.columns[2],
            df_kmerfinder.columns[4],
        ]
        for col in expected_columns:
            if col not in df_kmerfinder.columns:
                raise KeyError(f"Missing expected column: {col}")

        return {
            str(row["sample_name"]): {
                "colE": row.iloc[1],
                "colF": row.iloc[2],
                "colG": row.iloc[4],
            }
            for _, row in df_kmerfinder.iterrows()
        }

    except FileNotFoundError:
        print(f"Error: The file '{csv_file}' was not found.")
        return {}
    except pd.errors.EmptyDataError:
        print(f"Error: The file '{csv_file}' is empty.")
        return {}
    except pd.errors.ParserError:
        print(f"Error: The file '{csv_file}' could not be parsed.")
        return {}
    except KeyError as e:
        print(f"Error: {e}")
        return {}


def read_mapping_stats(mapping_file):
    """
    Reads a mapping statistics file and extracts mapping values per sample.

    Args:
        mapping_file (str): Path to the mapping statistics file.

    Returns:
        dict: A dictionary where keys are sample names and values are mapping statistics.
    """
    try:
        df_mapping = pd.read_csv(mapping_file, sep=";")
        df_mapping.columns = df_mapping.columns.str.strip()

        if (
            "SAMPLENAME" not in df_mapping.columns
            or "MAPPING" not in df_mapping.columns
        ):
            raise KeyError("Missing required columns: 'SAMPLENAME' or 'MAPPING'")

        return {
            str(row["SAMPLENAME"]): row["MAPPING"] for _, row in df_mapping.iterrows()
        }
    except FileNotFoundError:
        print(f"Error: The file '{mapping_file}' was not found.")
        return {}
    except pd.errors.EmptyDataError:
        print(f"Error: The file '{mapping_file}' is empty.")
        return {}
    except pd.errors.ParserError:
        print(f"Error: The file '{mapping_file}' could not be parsed.")
        return {}
    except KeyError as e:
        print(f"Error: {e}")
        return {}


def read_wgs_metrics(wgs_metrics_file):
    """
    Reads a Whole Genome Sequencing (WGS) metrics file and extracts relevant coverage metrics.

    Args:
        wgs_metrics_file (str): Path to the WGS metrics file.

    Returns:
        dict: A dictionary where keys are sample names and values are dictionaries containing
              mean coverage and percentage of reads above 10X coverage.
    """
    try:
        df_wgs = pd.read_csv(wgs_metrics_file, sep="\t", dtype={"SAMPLENAME": str})

        if (
            "SAMPLENAME" not in df_wgs.columns
            or "MEAN_COVERAGE" not in df_wgs.columns
            or "PCT_10X" not in df_wgs.columns
        ):
            raise KeyError(
                "Missing required columns: 'SAMPLENAME', 'MEAN_COVERAGE', or 'PCT_10X'"
            )

        return {
            str(row["SAMPLENAME"]): {
                "MEAN_COVERAGE": row["MEAN_COVERAGE"],
                "PCT_10X": row["PCT_10X"],
            }
            for _, row in df_wgs.iterrows()
        }
    except Exception as e:
        print(f"Error processing WGS metrics file: {e}")
        return {}


def read_virulence_stats(virulence_file):
    """
    Reads a virulence gene statistics file and processes the data.

    Args:
        virulence_file (str): Path to the virulence statistics file.

    Returns:
        dict: A dictionary where keys are sample names and values are dictionaries containing:
              - "genes": A string with detected virulence genes.
              - "count": The number of virulence genes detected.
    """
    df_virulence = pd.read_csv(virulence_file)

    virulence_dict = {}
    for _, row in df_virulence.iterrows():
        sample = str(row["sample"])

        # Process the virulence gene list, removing unnecessary characters
        virulence_genes = row["virulence"].strip("[]").replace("'", "").split(", ")
        virulence_genes_cleaned = [
            gene.replace(".match", "") for gene in virulence_genes
        ]
        virulence_list = ", ".join(virulence_genes_cleaned)

        # Get the virulence gene count
        virulence_count = row["virulence_genes_vfdb"]

        virulence_dict[sample] = {"genes": virulence_list, "count": virulence_count}

    return virulence_dict


def read_card_stats(card_file):
    """
    Reads a CARD (Comprehensive Antibiotic Resistance Database) file and extracts resistance gene data.

    Args:
        card_file (str): Path to the CARD resistance gene statistics file.

    Returns:
        dict: A dictionary where keys are sample names and values are dictionaries containing:
              - "genes": A string with detected resistance genes.
    """
    df_card = pd.read_csv(card_file)

    card_dict = {}
    for _, row in df_card.iterrows():
        sample = str(row["sample"])

        # Process the resistance gene list, removing unnecessary characters
        card_genes = (
            row["resistance_genes_car"].strip("[]").replace("'", "").split(", ")
        )
        card_genes_cleaned = [gene.replace(".match", "") for gene in card_genes]
        card_list = ", ".join(card_genes_cleaned)

        card_dict[sample] = {"genes": card_list}

    return card_dict


def read_amrfinder_results(directory):
    """
    Reads AMRFinder results from multiple TSV files in a given directory.

    Args:
        directory (str): Path to the directory containing AMRFinder result files.

    Returns:
        dict: A dictionary where keys are sample names and values are strings of detected resistance genes.
    """
    resistance_dict = {}

    for filename in os.listdir(directory):
        if filename.endswith("_out.tsv"):
            sample_id = filename.replace("_out.tsv", "")
            file_path = os.path.join(directory, filename)

            df = pd.read_csv(file_path, sep="\t")

            # Check if required columns exist
            if "Name" in df.columns and "Gene symbol" in df.columns:
                # Extract unique resistance genes
                filtered_genes = df["Gene symbol"].dropna().astype(str).unique()
                gene_list = ", ".join(filtered_genes)
                resistance_dict[sample_id] = gene_list

    return resistance_dict


def read_amrfinderplus_resistance(directory):
    """
    Reads AMRFinderPlus resistance data from multiple TSV files in a given directory.

    Args:
        directory (str): Path to the directory containing AMRFinderPlus result files.

    Returns:
        dict: A dictionary where keys are sample names and values are Pandas DataFrames with resistance data.
    """
    amr_data = {}

    for filename in os.listdir(directory):
        if filename.endswith("_out.tsv"):
            sample_id = filename.replace("_out.tsv", "")
            file_path = os.path.join(directory, filename)

            df = pd.read_csv(file_path, sep="\t")

            # Remove unnecessary columns
            columns_to_remove = [
                "Name",
                "Protein identifier",
                "HMM id",
                "HMM description",
            ]
            df_filtered = df.drop(
                columns=[col for col in columns_to_remove if col in df.columns]
            )

            # Store the filtered DataFrame in the dictionary
            amr_data[sample_id] = df_filtered

    return amr_data


def read_variants_stats(variants_stats_file):
    """
    Reads a variant statistics file and extracts SNP, deletion, insertion, and heterozygous mutation counts.

    Args:
        variants_stats_file (str): Path to the variant statistics file.

    Returns:
        dict: A dictionary where keys are sample names and values contain variant statistics.
    """
    try:
        df_variants = pd.read_csv(variants_stats_file, sep=";")
        df_variants.columns = df_variants.columns.str.strip()

        required_columns = ["SAMPLENAME", "SNP", "DEL", "INS", "HET"]
        for col in required_columns:
            if col not in df_variants.columns:
                raise KeyError(f"Missing required column: {col}")

        return {
            str(
                row["SAMPLENAME"]
            ): f"{row['SNP']};{row['DEL']};{row['INS']};{row['HET']}"
            for _, row in df_variants.iterrows()
        }
    except Exception as e:
        print(f"Error processing variants stats file: {e}")
        return {}


def read_quast_report(quast_file):
    """
    Reads a QUAST assembly report and extracts relevant quality metrics.

    Args:
        quast_file (str): Path to the QUAST report file.

    Returns:
        dict: A dictionary where keys are sample names and values contain assembly quality statistics.
    """
    try:
        df_quast = pd.read_csv(quast_file, sep="\t", header=0)
        required_columns = [
            "Assembly",
            "# contigs (>= 1000 bp)",
            "GC (%)",
            "L50",
            "N50",
            "Total length (>= 1000 bp)",
        ]
        for col in required_columns:
            if col not in df_quast.columns:
                raise KeyError(f"Missing required column: {col}")

        df_quast = df_quast.copy()
        df_quast.loc[:, "Sample"] = df_quast["Assembly"].str.replace(".scaffolds", "")

        return df_quast.set_index("Sample")[required_columns].to_dict(orient="index")
    except Exception as e:
        print(f"Error processing QUAST report file: {e}")
        return {}


def read_quast_per_reference(directory):
    """
    Reads all `transposed_report.tsv` files within subdirectories of the given directory.
    Extracts the `# genomic features` and `Genome fraction (%)` values per sample.

    Args:
        directory (str): Path to the QUAST per-reference reports directory.

    Returns:
        dict: A dictionary where keys are sample names (without `.scaffolds`)
              and values are a dictionary with:
              - "genomic_features": Number of genomic features
              - "genome_fraction": Genome fraction (%)
    """
    quast_data = {}

    # Recorrer todos los subdirectorios dentro del directorio dado
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file == "transposed_report.tsv":
                file_path = os.path.join(root, file)

                try:
                    df = pd.read_csv(file_path, sep="\t", header=0)

                    # Verificar que las columnas necesarias existen
                    required_columns = [
                        "Assembly",
                        "# genomic features",
                        "Genome fraction (%)",
                    ]
                    for col in required_columns:
                        if col not in df.columns:
                            raise KeyError(
                                f"Missing required column: {col} in {file_path}"
                            )

                    # Limpiar el nombre de la muestra eliminando `.scaffolds`
                    df["Sample"] = df["Assembly"].str.replace(
                        ".scaffolds", "", regex=False
                    )

                    # Crear diccionario de resultados
                    for _, row in df.iterrows():
                        sample = row["Sample"]
                        quast_data[sample] = {
                            "genomic_features": row["# genomic features"],
                            "genome_fraction": row["Genome fraction (%)"],
                        }

                except Exception as e:
                    print(f"Error processing QUAST report {file_path}: {e}")

    return quast_data


def read_plasmid_data(plasmid_file):
    """
    Reads a plasmid identification report and processes relevant columns.

    Args:
        plasmid_file (str): Path to the plasmid report file.

    Returns:
        DataFrame: A Pandas DataFrame containing plasmid information.
    """
    try:
        df = pd.read_csv(plasmid_file, sep="\t")

        # Column renaming dictionary
        column_rename = {
            "sample": "Sample",
            "id": "Plasmid ID",
            "length": "Length",
            "species": "Species",
            "description": "Description",
            "fraction_covered": "Fraction Covered",
            "contig_number": "Contig Number",
            "% Mapping": "% Mapping",
        }

        # Rename columns for clarity
        df.rename(columns=column_rename, inplace=True)
        return df
    except Exception as e:
        print(f"Error processing plasmid ID file: {e}")
        return {}


def read_mlva_results(mlva_file):
    """
    Reads the MLVA analysis CSV file and extracts relevant columns.

    Args:
        mlva_file (str): Path to the MLVA results file.

    Returns:
        tuple: (header list, dictionary where keys are sample IDs and values are lists of MLVA results).
    """
    try:
        df_mlva = pd.read_csv(mlva_file)

        if "Access_number" not in df_mlva.columns:
            raise KeyError("Missing required column: 'Access_number' in MLVA file")

        mlva_headers = df_mlva.columns[2:].tolist()

        mlva_dict = {
            str(row["Access_number"]): row.iloc[2:].tolist()
            for _, row in df_mlva.iterrows()
        }

        return mlva_headers, mlva_dict

    except Exception as e:
        print(f"Error reading MLVA file: {e}")
        return [], {}


# ------------------------------------------------------
# Excel Writing Functions
# ------------------------------------------------------


def add_samples(ws, samples):
    """
    Adds sample names to the first column of the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the samples will be added.
    samples (dict): Dictionary mapping sample names to their respective row numbers.
    """
    try:
        for sample, row_num in samples.items():
            ws[f"A{row_num}"] = sample
    except Exception as e:
        print(f"Error adding samples to xlsx: {e}")
        return {}


def add_ariba_mlst_stats(ws, samples, mlst_dict, mlst_header):
    """
    Adds ARIBA MLST (Multi-Locus Sequence Typing) results to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the results will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    mlst_dict (dict): Dictionary containing MLST results for each sample.
    mlst_header (str): Header label for the MLST column.
    """
    try:
        ws["C2"] = mlst_header
        for sample, row_num in samples.items():
            if sample in mlst_dict:
                ws[f"B{row_num}"] = mlst_dict[sample]["ST"]
                ws[f"C{row_num}"] = mlst_dict[sample]["genes"]
    except Exception as e:
        print(f"Error adding ariba mlst to xlsx: {e}")
        return {}


def add_kmerfinder_stats(ws, samples, kmerfinder_dict):
    """
    Adds KmerFinder results to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the results will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    kmerfinder_dict (dict): Dictionary containing KmerFinder results.
    """
    try:
        for sample, row_num in samples.items():
            if sample in kmerfinder_dict:
                ws[f"E{row_num}"] = kmerfinder_dict[sample]["colE"]
                ws[f"F{row_num}"] = kmerfinder_dict[sample]["colF"]
                ws[f"G{row_num}"] = kmerfinder_dict[sample]["colG"]
    except Exception as e:
        print(f"Error adding kmerfinder results to xlsx: {e}")
        return {}


def add_mapping_stats(ws, samples, mapping_dict, wgs_metrics_dict):
    """
    Adds mapping statistics to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the statistics will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    mapping_dict (dict): Dictionary containing mapping statistics.
    wgs_metrics_dict (dict): Dictionary containing whole-genome sequencing metrics.
    """
    try:
        for sample, row_num in samples.items():
            ws[f"H{row_num}"] = mapping_dict.get(sample, "NA")
            ws[f"I{row_num}"] = wgs_metrics_dict.get(sample, {}).get(
                "MEAN_COVERAGE", "NA"
            )
            ws[f"J{row_num}"] = wgs_metrics_dict.get(sample, {}).get("PCT_10X", "NA")
    except Exception as e:
        print(f"Error adding mapping stats to xlsx: {e}")
        return {}


def add_variants_stats(ws, samples, variants_dict):
    """
    Adds variant statistics to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the variant data will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    variants_dict (dict): Dictionary containing variant statistics.
    """
    try:
        for sample, row_num in samples.items():
            ws[f"K{row_num}"] = variants_dict.get(sample, "NA;NA;NA;NA")
    except Exception as e:
        print(f"Error adding variant stats to xlsx: {e}")
        return {}


def add_quast_stats(ws, samples, quast_dict):
    """
    Adds QUAST assembly quality statistics to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the QUAST statistics will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    quast_dict (dict): Dictionary containing QUAST results.
    """
    try:
        for sample, row_num in samples.items():
            sample_without_scaffolds = sample.replace(".scaffolds", "")
            if sample_without_scaffolds in quast_dict:
                quast_stats = quast_dict[sample_without_scaffolds]
                ws[f"L{row_num}"] = quast_stats.get("# contigs (>= 1000 bp)", "NA")
                ws[f"N{row_num}"] = quast_stats.get("GC (%)", "NA")
                ws[f"P{row_num}"] = quast_stats.get("L50", "NA")
                ws[f"Q{row_num}"] = quast_stats.get("N50", "NA")
                ws[f"R{row_num}"] = quast_stats.get("Total length (>= 1000 bp)", "NA")
    except Exception as e:
        print(f"Error adding quast stats to xlsx: {e}")
        return {}


def add_quast_per_reference(ws, samples, quast_dict):
    """
    Adds per-reference QUAST results to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the QUAST statistics will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    quast_dict (dict): Dictionary containing QUAST per-reference results.
    """
    try:
        for sample, row_num in samples.items():
            if sample in quast_dict:
                ws[f"M{row_num}"] = quast_dict[sample].get("genomic_features", "NA")
                ws[f"O{row_num}"] = quast_dict[sample].get("genome_fraction", "NA")
    except Exception as e:
        print(f"Error adding per-reference QUAST stats to xlsx: {e}")


def add_virulence_stats(ws, samples, virulence_dict):
    """
    Adds virulence gene statistics to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the virulence statistics will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    virulence_dict (dict): Dictionary containing virulence gene information.
    """
    try:
        for sample, row_num in samples.items():
            if sample in virulence_dict:
                ws[f"A{row_num-1}"] = sample
                ws[f"B{row_num-1}"] = virulence_dict[sample]["genes"]
                ws[f"C{row_num-1}"] = virulence_dict[sample]["count"]
    except Exception as e:
        print(f"Error adding virulence stats to xlsx: {e}")
        return {}


def add_resistance_stats(ws, samples, card_dict):
    """
    Adds antimicrobial resistance (AMR) gene statistics from CARD to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the AMR data will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    card_dict (dict): Dictionary containing AMR gene data.
    """
    try:
        for sample, row_num in samples.items():
            if sample in card_dict:
                ws[f"A{row_num-1}"] = sample
                ws[f"B{row_num-1}"] = card_dict[sample]["genes"]
    except Exception as e:
        print(f"Error adding resistance stats to xlsx: {e}")
        return {}


def add_amrfinder_results(ws, samples, resistance_dict):
    """
    Adds AMRFinder resistance gene results to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the AMRFinder results will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    resistance_dict (dict): Dictionary containing AMRFinder resistance gene results.
    """
    try:
        for sample, row_num in samples.items():
            if sample in resistance_dict:
                ws[f"C{row_num-1}"] = resistance_dict[sample]
    except Exception as e:
        print(f"Error adding amrfinder results to xlsx: {e}")
        return {}


def add_amrfinderplus_resistance(ws, samples, amr_data):
    """
    Adds AMRFinderPlus resistance gene results in a tabular format.

    Parameters:
    ws (Worksheet): The Excel worksheet where the AMRFinderPlus results will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    amr_data (dict): Dictionary containing AMRFinderPlus resistance gene data as Pandas DataFrames.
    """
    try:
        start_row = 2
        for sample, df in amr_data.items():
            if sample in samples:
                row_num = start_row
                for _, row in df.iterrows():
                    ws[f"A{row_num}"] = sample
                    for col_idx, value in enumerate(row.values, start=2):
                        ws.cell(row=row_num, column=col_idx, value=value)
                    row_num += 1
                start_row = row_num
    except Exception as e:
        print(f"Error adding amrfinder resistance stats to xlsx: {e}")
        return {}


def add_plasmid_data(ws, df):
    """
    Adds plasmid data from a Pandas DataFrame to the worksheet.

    Parameters:
    ws (Worksheet): The Excel worksheet where the plasmid data will be added.
    df (DataFrame): Pandas DataFrame containing plasmid information.
    """
    try:
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    except Exception as e:
        print(f"Error adding plasmidID stats to xlsx: {e}")
        return {}


def add_mlva_results(ws, mlva_headers, mlva_dict):
    """
    Adds MLVA results to the worksheet, creating new headers dynamically.

    Parameters:
    ws (Worksheet): The Excel worksheet where the MLVA results will be added.
    samples (dict): Dictionary mapping sample names to row numbers.
    mlva_headers (list): List of MLVA column headers.
    mlva_dict (dict): Dictionary containing MLVA results per sample.
    """
    try:
        ws.append(["Sample ID"] + mlva_headers)

        for sample_id, values in mlva_dict.items():
            ws.append([sample_id] + values)

    except Exception as e:
        print(f"Error adding MLVA results to xlsx: {e}")


# ------------------------------------------------------
# Main Functions
# ------------------------------------------------------


def main():
    samples = read_samples(samples_file)

    mlst_dict, mlst_header = read_ariba_mlst(tsv_file) if tsv_file else ({}, "")
    kmerfinder_dict = read_kmerfinder(csv_file) if csv_file else {}
    mapping_dict = read_mapping_stats(mapping_file) if mapping_file else {}
    wgs_metrics_dict = read_wgs_metrics(wgs_metrics_file) if wgs_metrics_file else {}
    variants_dict = (
        read_variants_stats(variants_stats_file) if variants_stats_file else {}
    )
    quast_dict = read_quast_report(quast_report_file) if quast_report_file else {}
    quast_per_reference_dict = read_quast_per_reference(quast_dir) if quast_dir else {}
    virulence_dict = read_virulence_stats(virulence_file) if virulence_file else {}
    resistance_dic = read_card_stats(card_file) if card_file else {}
    amrfinder_dict = read_amrfinder_results(amrfinder_dir) if amrfinder_dir else {}
    amr_resistance_data = (
        read_amrfinderplus_resistance(amrfinder_dir) if amrfinder_dir else {}
    )
    plasmid_data = read_plasmid_data(plasmid_file) if plasmid_file else pd.DataFrame()
    mlva_headers, mlva_dict = read_mlva_results(mlva_file) if mlva_file else ([], {})

    wb = load_workbook(xlsx_template)

    ws_summary = wb["summary"]
    ws_plasmids = wb["plasmids"]
    ws_virulence = wb["virulence"]
    ws_resistance = wb["Resistance result"]
    ws_mlva = wb["MLVA"]
    ws_amrfinder_resistance = wb["AMRFinderPlus Resistance result"]

    add_samples(ws_summary, samples)
    add_ariba_mlst_stats(ws_summary, samples, mlst_dict, mlst_header)
    add_kmerfinder_stats(ws_summary, samples, kmerfinder_dict)
    add_mapping_stats(ws_summary, samples, mapping_dict, wgs_metrics_dict)
    add_variants_stats(ws_summary, samples, variants_dict)
    add_quast_stats(ws_summary, samples, quast_dict)
    add_quast_per_reference(ws_summary, samples, quast_per_reference_dict)
    add_virulence_stats(ws_virulence, samples, virulence_dict)
    add_resistance_stats(ws_resistance, samples, resistance_dic)
    add_amrfinder_results(ws_resistance, samples, amrfinder_dict)
    add_amrfinderplus_resistance(ws_amrfinder_resistance, samples, amr_resistance_data)

    if not plasmid_data.empty:
        add_plasmid_data(ws_plasmids, plasmid_data)

    if mlva_headers and mlva_dict:
        add_mlva_results(ws_mlva, mlva_headers, mlva_dict)

    output_xlsx = "summary_outbreak_filled.xlsx"
    wb.save(output_xlsx)
    print(f"File saved: {output_xlsx}")


if __name__ == "__main__":
    main()
